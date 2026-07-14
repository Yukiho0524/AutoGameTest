"""Player-facing reports generated from autonomous exploration runs."""
from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = ROOT / "data" / "reports"
AUTONOMOUS_DIR = REPORTS_DIR / "autonomous"


def _safe_stem(value: str, default: str = "report") -> str:
    text = re.sub(r'[\\/:*?"<>|\s]+', "_", str(value or "").strip())
    text = re.sub(r"_+", "_", text).strip("._")
    return text[:80] or default


def _clip(value: Any, limit: int = 500) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text if len(text) <= limit else text[:limit - 3] + "..."


def _rel(path: Path) -> str:
    return path.resolve().relative_to(ROOT).as_posix()


def resolve_report_path(relative_path: str) -> Path | None:
    text = str(relative_path or "").replace("\\", "/").lstrip("/")
    if not text or ".." in Path(text).parts:
        return None
    full = (ROOT / text).resolve()
    try:
        full.relative_to(REPORTS_DIR.resolve())
    except ValueError:
        return None
    if not full.is_file() or full.suffix.lower() != ".xlsx":
        return None
    return full


def _turn_action(turn: dict) -> str:
    action = str(turn.get("action") or "").strip()
    detail = turn.get("execute_detail")
    return _clip(f"{action} / {detail}" if detail and detail != action else action, 160)


def _collect_player_feedback(turns: list[dict]) -> dict[str, list[str]]:
    positives: list[str] = []
    frictions: list[str] = []
    risks: list[str] = []
    suggestions: list[str] = []

    wait_count = 0
    back_count = 0
    error_count = 0
    explored_count = 0
    loading_hints = ("loading", "載入", "等待", "轉場", "黑畫面")
    risk_hints = ("登入", "授權", "付費", "購買", "抽卡", "轉蛋", "pvp", "排位", "匹配")

    for turn in turns:
        action = str(turn.get("action") or "").lower()
        status = str(turn.get("status") or "").lower()
        reason = _clip(turn.get("reason"), 260)
        observation = _clip(turn.get("observation"), 260)
        feeling = _clip(turn.get("player_feedback") or turn.get("feeling"), 260)
        learned = _clip(turn.get("learned"), 260)
        text = " ".join([reason, observation, feeling, learned]).lower()

        if action in {"tap", "swipe"} and status in {"done", "continue", ""}:
            explored_count += 1
            positives.append(
                feeling or observation or learned or f"第 {turn.get('turn')} 輪找到可互動入口，操作後可繼續探索。"
            )
        if action == "wait" or any(h.lower() in text for h in loading_hints):
            wait_count += 1
            frictions.append(
                feeling or observation or reason or f"第 {turn.get('turn')} 輪主要在等待載入或轉場。"
            )
        if action == "back":
            back_count += 1
            frictions.append(
                feeling or observation or reason or f"第 {turn.get('turn')} 輪需要返回上一層，表示目前路徑不夠明確。"
            )
        if status in {"error", "stopped", "timeout"}:
            error_count += 1
            frictions.append(feeling or reason or f"第 {turn.get('turn')} 輪探索中斷或達到限制。")
        if any(h.lower() in text for h in risk_hints):
            risks.append(feeling or observation or reason or learned)

    if explored_count:
        suggestions.append("保留目前可辨識的入口與導覽提示，玩家能靠畫面元素逐步探索。")
    if wait_count:
        suggestions.append("載入與轉場處建議提供更清楚的進度、下一步提示或可點擊狀態，降低等待時的不確定感。")
    if back_count:
        suggestions.append("容易走到非目標頁面時，可強化返回、麵包屑或主功能入口，讓玩家更快回到核心流程。")
    if error_count:
        suggestions.append("探索中斷或不確定畫面建議增加明確文字、教學提示或安全退出方式。")
    if risks:
        suggestions.append("登入、付費、抽卡或 PVP 相關入口需維持清楚標示與確認流程，避免玩家誤觸。")
    if not suggestions:
        suggestions.append("本次探索未收集到明顯卡點；可延長自主探索時間，觀察更深層功能流程。")

    return {
        "positives": _unique(positives, 8),
        "frictions": _unique(frictions, 10),
        "risks": _unique(risks, 8),
        "suggestions": _unique(suggestions, 8),
    }


def _unique(items: list[str], limit: int) -> list[str]:
    rows: list[str] = []
    for item in items:
        text = _clip(item, 360).strip(" -")
        if text and text not in rows:
            rows.append(text)
        if len(rows) >= limit:
            break
    return rows


def _append_section(ws, title: str, rows: list[str], header_font, wrap) -> None:
    ws.append([])
    ws.append([title])
    ws.cell(ws.max_row, 1).font = header_font
    if not rows:
        ws.append(["未收集到明確內容。"])
        ws.cell(ws.max_row, 1).alignment = wrap
        return
    for index, text in enumerate(rows, 1):
        ws.append([f"{index}. {text}"])
        ws.cell(ws.max_row, 1).alignment = wrap


def write_autonomous_player_report(game: dict, job_id: str | None,
                                   result: dict, performance: dict,
                                   base_dir: Path | None = None) -> dict:
    """Create an xlsx report from autonomous exploration observations."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，無法寫出自主探索玩家回饋報告") from e

    turns = result.get("visual_turns") or performance.get("visual_turns") or []
    if not isinstance(turns, list):
        turns = []
    generated_at = datetime.now()
    out_dir = Path(base_dir) if base_dir else AUTONOMOUS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    game_id = _safe_stem(game.get("id") or "game")
    game_name = str(game.get("name") or game_id)
    job_text = _safe_stem(job_id or "manual")
    out = out_dir / f"{game_id}_{job_text}_player_feedback_{generated_at:%Y%m%d_%H%M%S}.xlsx"

    feedback = _collect_player_feedback(turns)
    wb = Workbook()
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", start_color="D9EAD3")
    wrap = Alignment(vertical="top", wrap_text=True)

    ws = wb.active
    ws.title = "玩家體驗摘要"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 96
    ws["A1"] = "自主探索玩家體驗回饋"
    ws["A1"].font = title_font
    meta_rows = [
        ("遊戲", game_name),
        ("game_id", game.get("id", "")),
        ("job_id", job_id or "manual"),
        ("產出時間", generated_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("探索輪數", len(turns)),
        ("總耗時秒數", performance.get("total_seconds", "")),
        ("結束原因", result.get("reason", "")),
        ("玩家視角總評", _overall_comment(turns, feedback, result)),
    ]
    for row in meta_rows:
        ws.append(row)
        ws.cell(ws.max_row, 1).font = header_font
        ws.cell(ws.max_row, 2).alignment = wrap
    _append_section(ws, "覺得順暢 / 有吸引力的地方", feedback["positives"], header_font, wrap)
    _append_section(ws, "玩家可能卡住或困惑的地方", feedback["frictions"], header_font, wrap)
    _append_section(ws, "需要清楚標示的風險入口", feedback["risks"], header_font, wrap)
    _append_section(ws, "改善建議", feedback["suggestions"], header_font, wrap)

    ws_turns = wb.create_sheet("逐輪體驗紀錄")
    headers = ["輪次", "狀態", "動作", "玩家觀察", "玩家感受", "學到的事", "下一狀態", "原因", "截圖"]
    ws_turns.append(headers)
    for col in range(1, len(headers) + 1):
        ws_turns.cell(1, col).font = header_font
        ws_turns.cell(1, col).fill = header_fill
    widths = {"A": 7, "B": 12, "C": 18, "D": 42, "E": 42, "F": 42, "G": 34, "H": 48, "I": 56}
    for col, width in widths.items():
        ws_turns.column_dimensions[col].width = width
    for turn in turns:
        ws_turns.append([
            turn.get("turn", ""),
            turn.get("status", ""),
            _turn_action(turn),
            _clip(turn.get("observation"), 700),
            _clip(turn.get("player_feedback") or turn.get("feeling") or turn.get("reason"), 700),
            _clip(turn.get("learned"), 700),
            _clip(turn.get("next_state"), 500),
            _clip(turn.get("reason"), 700),
            turn.get("screenshot", ""),
        ])
        for col in range(1, len(headers) + 1):
            ws_turns.cell(ws_turns.max_row, col).alignment = wrap
    ws_turns.freeze_panes = "A2"
    ws_turns.auto_filter.ref = ws_turns.dimensions

    ws_meta = wb.create_sheet("AutoGameTest")
    ws_meta.sheet_state = "hidden"
    for key, value in {
        "game_id": game.get("id", ""),
        "game_name": game_name,
        "job_id": job_id or "manual",
        "report_kind": "autonomous_player_feedback",
        "generated_at": generated_at.strftime("%Y-%m-%d %H:%M:%S"),
        "relative_path": "",
    }.items():
        ws_meta.append([key, value])

    try:
        wb.save(out)
    except PermissionError:
        out = out_dir / f"{game_id}_{job_text}_player_feedback_{generated_at:%Y%m%d_%H%M%S}_{os.getpid()}.xlsx"
        wb.save(out)

    rel = _rel(out)
    ws_meta["B6"] = rel
    wb.save(out)
    return {
        "ok": True,
        "path": str(out),
        "relative_path": rel,
        "name": out.name,
        "turns": len(turns),
        "generated_at": generated_at.strftime("%Y-%m-%d %H:%M:%S"),
    }


def _overall_comment(turns: list[dict], feedback: dict[str, list[str]], result: dict) -> str:
    if not turns:
        return "本次沒有取得逐輪探索紀錄，暫時無法形成玩家體驗判斷。"
    if feedback.get("frictions"):
        return "以玩家角度來看，目前已能探索部分入口，但仍存在載入、導覽或不確定畫面造成的中斷感。"
    if feedback.get("positives"):
        return "以玩家角度來看，畫面入口具備可探索性，能透過低風險操作逐步理解主要功能。"
    reason = _clip(result.get("reason"), 220)
    return f"本次探索完成，整體體驗需搭配更長時間觀察；結束原因：{reason}" if reason else "本次探索完成，建議延長時間觀察更深層流程。"
