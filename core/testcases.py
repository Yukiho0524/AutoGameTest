"""Generate QA TestCase workbooks from planning documents."""
from __future__ import annotations

import base64
import hashlib
import html
import json
import os
import re
import shutil
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
TESTCASE_DIR = ROOT / "TestCase"
INPUT_DIR = TESTCASE_DIR / "_input"
UPLOAD_DIR = INPUT_DIR / "uploads"
SPEC_PATH = ROOT / "TESTCASE_SPEC.md"
SYSTEM_SKILLS_DIR = ROOT / "data" / "system_skills"
SUPPORTED_EXTS = {".docx", ".pdf", ".xlsx", ".xlsm", ".txt", ".md"}
VALID_TYPES = {"顯示確認", "操作確認", "數值確認"}
VALID_PRIORITIES = {"P0", "P1", "P2"}
VALID_AUTOMATION = {"A", "S", "M"}
VALID_DESTRUCTIVE_RISKS = {"SAFE", "GUARDED", "MANUAL"}
VALID_DESTRUCTIVE_TYPES = {
    "連點",
    "邊界",
    "狀態中斷",
    "流程逆行",
    "資源不足",
    "斷線重連",
    "重啟恢復",
    "高風險人工",
}


def _safe_stem(value: str) -> str:
    stem = Path(value or "企劃書").stem
    return re.sub(r'[\\/:*?"<>|]+', "_", stem).strip() or "企劃書"


def _safe_name(value: str) -> str:
    name = Path(value or "planning.txt").name
    name = re.sub(r'[\\/:*?"<>|]+', "_", name).strip()
    return name or "planning.txt"


def safe_name(value: str) -> str:
    return _safe_name(value)


def _clip(text: str, limit: int = 120_000) -> str:
    text = text or ""
    if len(text) <= limit:
        return text
    return text[:limit] + f"\n\n[內容過長，已截斷 {len(text) - limit} 字元]"


def ensure_dirs() -> None:
    TESTCASE_DIR.mkdir(parents=True, exist_ok=True)
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def save_upload(filename: str, content_base64: str) -> Path:
    ensure_dirs()
    safe = _safe_name(filename)
    ext = Path(safe).suffix.lower()
    if ext not in SUPPORTED_EXTS:
        raise ValueError(f"不支援的企劃書格式：{ext or '(無副檔名)'}")
    data = base64.b64decode(content_base64 or "", validate=True)
    if not data:
        raise ValueError("上傳檔案是空的")
    out = UPLOAD_DIR / f"{datetime.now():%Y%m%d_%H%M%S}_{safe}"
    out.write_bytes(data)
    return out


def extract_doc_text(path: str | os.PathLike) -> tuple[str, str]:
    """Return (text, mode)."""
    doc = Path(path)
    ext = doc.suffix.lower()
    if ext in (".txt", ".md"):
        return doc.read_text(encoding="utf-8", errors="replace"), "text"
    if ext == ".docx":
        return _extract_docx_text(doc), "docx-text"
    if ext in (".xlsx", ".xlsm"):
        return _extract_xlsx_text(doc), "xlsx-text"
    if ext == ".pdf":
        text = _extract_pdf_text(doc)
        if text.strip():
            return text, "pdf-text"
        return (
            "無法抽取 PDF 文字。請安裝 pdftotext，或改上傳 txt/md/docx/xlsx，"
            "或先將企劃書另存成可複製文字的 PDF。",
            "pdf-unreadable",
        )
    try:
        return doc.read_text(encoding="utf-8", errors="replace"), "text"
    except OSError:
        return "", "unknown"


def _extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", "replace")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"</w:tr>", "\n", xml)
    xml = re.sub(r"</w:tc>", "\t", xml)
    return html.unescape(re.sub(r"<[^>]+>", "", xml))


def _extract_xlsx_text(path: Path) -> str:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，無法讀取 xlsx 或寫出 TestCase xlsx") from e
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    lines: list[str] = []
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            lines.append(f"===== 工作表：{sheet} =====")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    lines.append("\t".join(cells))
    finally:
        wb.close()
    return "\n".join(lines)


def _extract_pdf_text(path: Path) -> str:
    exe = shutil.which("pdftotext")
    if not exe:
        return ""
    try:
        proc = subprocess.run(
            [exe, "-layout", "-enc", "UTF-8", str(path), "-"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
    except (OSError, subprocess.SubprocessError):
        return ""
    return proc.stdout or ""


def build_prompt(doc_name: str, doc_text: str, mode: str) -> str:
    spec = SPEC_PATH.read_text(encoding="utf-8", errors="replace")
    return f"""你是遊戲 QA 測試設計專家。請根據「測試案例撰寫規範」與企劃書內容，產出 QA TestCase。

# 測試案例撰寫規範
{spec}

# 企劃書
- 檔名：{doc_name}
- 讀取模式：{mode}

```text
{_clip(doc_text)}
```

# 重要原則
- 只根據企劃書明確文字撰寫 CASE。
- 企劃書沒寫清楚、PDF 無法讀到、圖示看不到的內容，一律列成 ISSUE。
- 不要補充企劃書沒有的數值、流程、UI 或常識推測。
- 每條 CASE 要可執行、可觀察、可判定。
- 另外輸出 3 到 8 條 SKILL，描述該系統的用途、入口、核心流程、狀態、限制或風險；仍然只能根據企劃書文字。

# 輸出格式
只輸出以下純文字行，不要輸出 Markdown 表格或額外說明：

SYSTEM: <系統名稱>
SKILL: <給 Agent 的系統理解重點>
CASE: <項目>|<類型>|<優先度>|<自動化>|<TC內容>
ISSUE: <待釐清問題>
"""


def parse_output(output: str) -> dict[str, Any]:
    system = ""
    skill_notes: list[str] = []
    cases: list[tuple[str, str, str, str, str]] = []
    issues: list[str] = []
    for line in str(output or "").splitlines():
        text = line.strip()
        text = re.sub(r"^[-*]\s*", "", text)
        m = re.match(r"^SYSTEM:\s*(.+)$", text)
        if m:
            system = m.group(1).strip()
            continue
        m = re.match(r"^SKILL:\s*(.+)$", text)
        if m:
            note = m.group(1).strip()
            if note:
                skill_notes.append(note)
            continue
        m = re.match(r"^CASE:\s*(.+)$", text)
        if m:
            parts = [p.strip() for p in m.group(1).split("|")]
            if len(parts) >= 5:
                item, typ, priority, automation = parts[:4]
                tc = "|".join(parts[4:]).strip()
                if typ not in VALID_TYPES:
                    typ = "操作確認"
                if priority not in VALID_PRIORITIES:
                    priority = "P1"
                if automation not in VALID_AUTOMATION:
                    automation = "M"
                if item and tc:
                    cases.append((item, typ, priority, automation, tc))
            continue
        m = re.match(r"^ISSUE:\s*(.+)$", text)
        if m:
            issues.append(m.group(1).strip())
    return {
        "system": system,
        "skill_notes": skill_notes,
        "cases": cases,
        "issues": issues,
    }


def build_destructive_prompt(meta: dict[str, Any], system_skill: str = "") -> str:
    source_cases = "\n".join(
        f"- TC{case['no']:03d}｜{case.get('item') or '未分類'}｜"
        f"{case.get('type') or '未分類'}｜{case.get('tc') or ''}"
        for case in meta.get("selected", [])
    )
    return f"""你是資深遊戲 QA，請根據既有 TestCase 與系統理解，產出「破壞性測試」案例。

# 目標
- 不是正常流程測試，而是負向、邊界、異常、狀態中斷、重複操作、資源不足、重啟/斷線恢復等測試。
- 每條案例必須可觀察、可判定，且要有風險分級。
- 不可要求 AI 執行登入、付費、購買確認、抽卡確認、刪帳號、第三方授權、PVP 或不可逆操作。

# 風險分級
- SAFE：可由 AI 自動執行的低風險測試，例如返回、切頁、連點不消耗的按鈕、點不可用按鈕、取消確認視窗。
- GUARDED：需要測試帳號、可回復狀態或人工確認測試環境，例如免費資源消耗、斷線重連、清空暫存、重啟恢復。
- MANUAL：只產生測項，不可由 AI 自動執行，例如真實購買、抽卡確認、刪帳號、帳號綁定、PVP、不可逆資料破壞。

# 既有 TestCase
TestCase 文件：{meta.get('name', '')}
遊戲：{meta.get('game_name', '') or meta.get('game_id', '')}
系統：{meta.get('system', '')}
來源企劃書：{meta.get('source_doc', '')}

```text
{_clip(source_cases, 40_000)}
```

# 系統理解 Skill
```text
{_clip(system_skill, 30_000)}
```

# 輸出格式
只輸出以下純文字行，不要輸出 Markdown 表格或額外說明：

SYSTEM: <系統名稱>
DCASE: <項目>|<風險 SAFE/GUARDED/MANUAL>|<破壞類型>|<優先度>|<自動化>|<TC內容>
ISSUE: <待釐清問題>
"""


def parse_destructive_output(output: str) -> dict[str, Any]:
    system = ""
    cases: list[tuple[str, str, str, str, str, str]] = []
    issues: list[str] = []
    for line in str(output or "").splitlines():
        text = re.sub(r"^[-*]\s*", "", line.strip())
        m = re.match(r"^SYSTEM:\s*(.+)$", text)
        if m:
            system = m.group(1).strip()
            continue
        m = re.match(r"^DCASE:\s*(.+)$", text)
        if m:
            parts = [p.strip() for p in m.group(1).split("|")]
            if len(parts) >= 6:
                item, risk, destructive_type, priority, automation = parts[:5]
                tc = "|".join(parts[5:]).strip()
                risk = risk.upper()
                if risk not in VALID_DESTRUCTIVE_RISKS:
                    risk = "MANUAL"
                if destructive_type not in VALID_DESTRUCTIVE_TYPES:
                    destructive_type = "高風險人工" if risk == "MANUAL" else "邊界"
                if priority not in VALID_PRIORITIES:
                    priority = "P1"
                if automation not in VALID_AUTOMATION:
                    automation = "M"
                if risk == "MANUAL":
                    automation = "M"
                if item and tc:
                    cases.append((item, risk, destructive_type, priority, automation, tc))
            continue
        m = re.match(r"^ISSUE:\s*(.+)$", text)
        if m:
            issues.append(m.group(1).strip())
    return {"system": system, "cases": cases, "issues": issues}


def write_testcase_xlsx(system: str, cases: list[tuple],
                        issues: list[str], doc_name: str,
                        game: dict | None = None,
                        system_skill: dict | None = None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，無法寫出 TestCase xlsx") from e

    ensure_dirs()
    out = TESTCASE_DIR / f"{_safe_stem(doc_name)}_TestCase.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = (re.sub(r'[\\/:*?"<>|]+', "_", system or "系統") or "系統")[:31]
    header = ["項目", "類型", "TC內容", "PASS/FAIL", "Bug Key", "備註",
              "PASS", "FAIL", "N/A", "Total", "TC數量"]
    ws.append(header)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    for col in range(1, len(header) + 1):
        ws.cell(1, col).font = header_font
        ws.cell(1, col).fill = header_fill
    for col, width in {"A": 18, "B": 12, "C": 86, "D": 11, "E": 12,
                       "F": 24, "G": 7, "H": 7, "I": 7, "J": 7, "K": 9}.items():
        ws.column_dimensions[col].width = width
    wrap = Alignment(vertical="top", wrap_text=True)
    prev_item = None
    for item, typ, _priority, _automation, tc in cases:
        ws.append([item if item != prev_item else "", typ, tc, "", "", ""])
        ws.cell(ws.max_row, 3).alignment = wrap
        prev_item = item
    ws["G2"] = '=COUNTIF(D:D,"PASS")'
    ws["H2"] = '=COUNTIF(D:D,"FAIL")'
    ws["I2"] = '=COUNTIF(D:D,"N/A")'
    ws["J2"] = "=G2+H2+I2"
    ws["K2"] = "=COUNTA(C2:C10000)"

    ws_meta = wb.create_sheet("優先度與自動化")
    ws_meta.append(["#", "項目", "類型", "優先度", "自動化(A=AI白話/S=腳本回歸/M=手動)", "TC內容"])
    for col in range(1, 7):
        ws_meta.cell(1, col).font = header_font
    for col, width in {"A": 5, "B": 18, "C": 12, "D": 8, "E": 34, "F": 86}.items():
        ws_meta.column_dimensions[col].width = width
    for index, (item, typ, priority, automation, tc) in enumerate(cases, 1):
        ws_meta.append([index, item, typ, priority, automation, tc])
        ws_meta.cell(ws_meta.max_row, 6).alignment = wrap

    ws_issue = wb.create_sheet("待釐清問題")
    ws_issue.column_dimensions["A"].width = 120
    ws_issue.append([f"企劃書：{doc_name}"])
    ws_issue.append([f"產出時間：{datetime.now():%Y-%m-%d %H:%M}"])
    ws_issue.append([f"案例數：{len(cases)}"])
    ws_issue.append([])
    ws_issue.append(["待釐清問題"])
    ws_issue.cell(ws_issue.max_row, 1).font = header_font
    for index, issue in enumerate(issues, 1):
        ws_issue.append([f"{index}. {issue}"])
        ws_issue.cell(ws_issue.max_row, 1).alignment = wrap

    ws_agt = wb.create_sheet("AutoGameTest")
    ws_agt.sheet_state = "hidden"
    ws_agt.append(["key", "value"])
    ws_agt.append(["game_id", (game or {}).get("id", "")])
    ws_agt.append(["game_name", (game or {}).get("name", "")])
    ws_agt.append(["source_doc", doc_name])
    ws_agt.append(["system", system or ""])
    ws_agt.append(["testcase_kind", "standard"])
    ws_agt.append(["system_skill_name", (system_skill or {}).get("name", "")])
    ws_agt.append(["system_skill_path", (system_skill or {}).get("relative_path", "")])
    ws_agt.append(["generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    try:
        wb.save(out)
    except PermissionError:
        out = TESTCASE_DIR / f"{_safe_stem(doc_name)}_TestCase_{datetime.now():%H%M%S}.xlsx"
        wb.save(out)
    return out


def _destructive_output_name(source_name: str) -> str:
    stem = _safe_stem(source_name)
    stem = re.sub(r"(_TestCase|_DestructiveTestCase)$", "", stem)
    return f"{stem}_DestructiveTestCase.xlsx"


def write_destructive_xlsx(system: str, cases: list[tuple],
                           issues: list[str], source_name: str,
                           base_meta: dict[str, Any]) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，無法寫出破壞性 TestCase xlsx") from e

    ensure_dirs()
    out = TESTCASE_DIR / _destructive_output_name(source_name)
    wb = Workbook()
    ws = wb.active
    ws.title = (re.sub(r'[\\/:*?"<>|]+', "_", system or "破壞性測試") or "破壞性測試")[:31]
    header = ["項目", "風險", "破壞類型", "TC內容", "PASS/FAIL", "Bug Key", "備註",
              "PASS", "FAIL", "N/A", "Total", "TC數量"]
    ws.append(header)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", start_color="FCE4D6")
    for col in range(1, len(header) + 1):
        ws.cell(1, col).font = header_font
        ws.cell(1, col).fill = header_fill
    widths = {
        "A": 18, "B": 12, "C": 14, "D": 86, "E": 11, "F": 12,
        "G": 24, "H": 7, "I": 7, "J": 7, "K": 7, "L": 9,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    wrap = Alignment(vertical="top", wrap_text=True)
    prev_item = None
    risk_counts = {"SAFE": 0, "GUARDED": 0, "MANUAL": 0}
    for item, risk, destructive_type, _priority, _automation, tc in cases:
        risk_counts[risk] = risk_counts.get(risk, 0) + 1
        ws.append([item if item != prev_item else "", risk, destructive_type, tc, "", "", ""])
        ws.cell(ws.max_row, 4).alignment = wrap
        prev_item = item
    ws["H2"] = '=COUNTIF(E:E,"PASS")'
    ws["I2"] = '=COUNTIF(E:E,"FAIL")'
    ws["J2"] = '=COUNTIF(E:E,"N/A")'
    ws["K2"] = "=H2+I2+J2"
    ws["L2"] = "=COUNTA(D2:D10000)"

    ws_meta = wb.create_sheet("風險與自動化")
    ws_meta.append(["#", "項目", "風險", "破壞類型", "優先度", "自動化", "TC內容"])
    for col in range(1, 8):
        ws_meta.cell(1, col).font = header_font
    for col, width in {"A": 5, "B": 18, "C": 12, "D": 14, "E": 8, "F": 10, "G": 86}.items():
        ws_meta.column_dimensions[col].width = width
    for index, (item, risk, destructive_type, priority, automation, tc) in enumerate(cases, 1):
        ws_meta.append([index, item, risk, destructive_type, priority, automation, tc])
        ws_meta.cell(ws_meta.max_row, 7).alignment = wrap

    ws_issue = wb.create_sheet("待釐清問題")
    ws_issue.column_dimensions["A"].width = 120
    ws_issue.append([f"來源 TestCase：{source_name}"])
    ws_issue.append([f"產出時間：{datetime.now():%Y-%m-%d %H:%M}"])
    ws_issue.append([f"SAFE：{risk_counts.get('SAFE', 0)}"])
    ws_issue.append([f"GUARDED：{risk_counts.get('GUARDED', 0)}"])
    ws_issue.append([f"MANUAL：{risk_counts.get('MANUAL', 0)}"])
    ws_issue.append([])
    ws_issue.append(["待釐清問題"])
    ws_issue.cell(ws_issue.max_row, 1).font = header_font
    for index, issue in enumerate(issues, 1):
        ws_issue.append([f"{index}. {issue}"])
        ws_issue.cell(ws_issue.max_row, 1).alignment = wrap

    ws_agt = wb.create_sheet("AutoGameTest")
    ws_agt.sheet_state = "hidden"
    ws_agt.append(["key", "value"])
    ws_agt.append(["game_id", base_meta.get("game_id", "")])
    ws_agt.append(["game_name", base_meta.get("game_name", "")])
    ws_agt.append(["source_doc", base_meta.get("source_doc", "")])
    ws_agt.append(["source_testcase", source_name])
    ws_agt.append(["system", system or ""])
    ws_agt.append(["testcase_kind", "destructive"])
    ws_agt.append(["risk_safe", str(risk_counts.get("SAFE", 0))])
    ws_agt.append(["risk_guarded", str(risk_counts.get("GUARDED", 0))])
    ws_agt.append(["risk_manual", str(risk_counts.get("MANUAL", 0))])
    ws_agt.append(["system_skill_name", base_meta.get("system_skill_name", "")])
    ws_agt.append(["system_skill_path", base_meta.get("system_skill_path", "")])
    ws_agt.append(["generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    try:
        wb.save(out)
    except PermissionError:
        out = TESTCASE_DIR / f"{_safe_stem(source_name)}_DestructiveTestCase_{datetime.now():%H%M%S}.xlsx"
        wb.save(out)
    return out


def _read_xlsx_metadata(path: Path) -> dict[str, str]:
    try:
        import openpyxl
    except ImportError:
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return {}
    try:
        if "AutoGameTest" not in wb.sheetnames:
            return {}
        ws = wb["AutoGameTest"]
        data = {}
        for key, value, *_ in ws.iter_rows(min_row=2, values_only=True):
            if key:
                data[str(key)] = "" if value is None else str(value)
        return data
    finally:
        wb.close()


def _iter_testcase_files() -> list[Path]:
    if not TESTCASE_DIR.exists():
        return []
    paths = set(TESTCASE_DIR.glob("*_TestCase*.xlsx"))
    paths.update(TESTCASE_DIR.glob("*_DestructiveTestCase*.xlsx"))
    return sorted(paths, key=lambda p: p.stat().st_mtime, reverse=True)


def list_testcases() -> list[dict[str, Any]]:
    if not TESTCASE_DIR.exists():
        return []
    rows = []
    for path in _iter_testcase_files():
        meta = _read_xlsx_metadata(path)
        rows.append({
            "name": path.name,
            "path": str(path),
            "size": path.stat().st_size,
            "mtime": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            "game_id": meta.get("game_id", ""),
            "game_name": meta.get("game_name", ""),
            "source_doc": meta.get("source_doc", ""),
            "source_testcase": meta.get("source_testcase", ""),
            "system": meta.get("system", ""),
            "testcase_kind": meta.get("testcase_kind", "standard"),
            "risk_safe": int(meta.get("risk_safe", "0") or 0),
            "risk_guarded": int(meta.get("risk_guarded", "0") or 0),
            "risk_manual": int(meta.get("risk_manual", "0") or 0),
            "system_skill_name": meta.get("system_skill_name", ""),
            "system_skill_path": meta.get("system_skill_path", ""),
        })
    return rows


def testcase_path(name: str) -> Path | None:
    safe = _safe_name(name)
    path = TESTCASE_DIR / safe
    if path.is_file() and path.suffix.lower() == ".xlsx":
        return path
    return None


def _resolve_project_path(value: str) -> Path | None:
    text = str(value or "").strip()
    if not text:
        return None
    path = Path(text)
    if not path.is_absolute():
        path = ROOT / path
    try:
        return path.resolve()
    except OSError:
        return None


def _path_is_under(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root.resolve())
        return True
    except (OSError, ValueError):
        return False


def _referenced_system_skill_paths(exclude_testcase: Path | None = None) -> set[str]:
    refs: set[str] = set()
    if not TESTCASE_DIR.exists():
        return refs
    exclude_resolved = exclude_testcase.resolve() if exclude_testcase else None
    for path in _iter_testcase_files():
        try:
            if exclude_resolved and path.resolve() == exclude_resolved:
                continue
        except OSError:
            continue
        meta = _read_xlsx_metadata(path)
        ref = _resolve_project_path(meta.get("system_skill_path", ""))
        if ref:
            refs.add(str(ref).lower())
    return refs


def _remove_system_skill_index_entry(skill_path: Path, skill_name: str) -> bool:
    index_path = skill_path.parent.parent / "index.json"
    if not _path_is_under(index_path, SYSTEM_SKILLS_DIR) or not index_path.is_file():
        return False
    try:
        data = json.loads(index_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False
    systems = data.get("systems")
    if not isinstance(systems, dict):
        return False
    key = skill_name or skill_path.parent.name
    if key not in systems:
        return False
    systems.pop(key, None)
    data["systems"] = dict(sorted(systems.items()))
    data["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    index_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return True


def _cleanup_orphan_system_skill(meta: dict[str, str],
                                 deleted_testcase: Path) -> dict[str, Any]:
    skill_path = _resolve_project_path(meta.get("system_skill_path", ""))
    if not skill_path:
        return {"deleted": False, "reason": "no system skill"}
    if not _path_is_under(skill_path, SYSTEM_SKILLS_DIR):
        return {"deleted": False, "reason": "outside system skill dir"}
    if skill_path.name != "SKILL.md":
        return {"deleted": False, "reason": "not a generated SKILL.md"}
    if str(skill_path).lower() in _referenced_system_skill_paths(deleted_testcase):
        return {"deleted": False, "reason": "still referenced"}

    index_updated = _remove_system_skill_index_entry(
        skill_path,
        meta.get("system_skill_name", ""),
    )
    try:
        if skill_path.parent.is_dir():
            shutil.rmtree(skill_path.parent)
        return {
            "deleted": True,
            "path": str(skill_path),
            "index_updated": index_updated,
        }
    except OSError as e:
        return {"deleted": False, "reason": str(e), "index_updated": index_updated}


def delete_testcase(name: str) -> dict[str, Any]:
    path = testcase_path(name)
    if not path:
        return {"ok": False, "error": f"找不到 TestCase 文件：{name}"}
    meta = _read_xlsx_metadata(path)
    try:
        path.unlink()
    except OSError as e:
        return {"ok": False, "error": f"刪除失敗：{e}"}
    cleanup = _cleanup_orphan_system_skill(meta, path)
    return {
        "ok": True,
        "deleted": str(path),
        "system_skill": cleanup,
    }


def _normalize_case_limit(limit: Any, default: int = 25) -> int | None:
    text = str(limit if limit is not None else "").strip().lower()
    if text in {"all", "全部"}:
        return None
    try:
        value = int(limit)
    except (TypeError, ValueError):
        value = default
    if value <= 0:
        return None
    return max(1, min(value, 80))


def _slice_cases(cases: list[dict[str, Any]], limit: int | None) -> list[dict[str, Any]]:
    return list(cases) if limit is None else list(cases[:limit])


def read_testcase_cases(name: str, limit: int | str | None = 25) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，無法讀取 TestCase xlsx") from e

    path = testcase_path(name)
    if not path:
        raise FileNotFoundError(f"找不到 TestCase 文件：{name}")
    normalized_limit = _normalize_case_limit(limit)
    metadata = _read_xlsx_metadata(path)
    testcase_kind = metadata.get("testcase_kind", "standard") or "standard"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        system = ws.title
        cases: list[dict[str, Any]] = []
        current_item = ""
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cells = list(row or []) + [None] * 8
            if cells[0]:
                current_item = str(cells[0]).strip()
            if testcase_kind == "destructive":
                tc_text = str(cells[3] or "").strip()
                if not tc_text:
                    continue
                risk = str(cells[1] or "").strip().upper()
                if risk not in VALID_DESTRUCTIVE_RISKS:
                    risk = "MANUAL"
                destructive_type = str(cells[2] or "").strip() or "邊界"
                cases.append({
                    "no": len(cases) + 1,
                    "row": row_index,
                    "item": current_item,
                    "risk": risk,
                    "type": destructive_type,
                    "tc": tc_text,
                    "result": str(cells[4] or "").strip(),
                })
            else:
                tc_text = str(cells[2] or "").strip()
                if not tc_text:
                    continue
                cases.append({
                    "no": len(cases) + 1,
                    "row": row_index,
                    "item": current_item,
                    "type": str(cells[1] or "").strip(),
                    "tc": tc_text,
                    "result": str(cells[3] or "").strip(),
                })
    finally:
        wb.close()
    if testcase_kind == "destructive":
        safe_cases = [case for case in cases if case.get("risk") == "SAFE"]
        pending = [case for case in safe_cases if not case.get("result")]
        selected = _slice_cases(pending or safe_cases, normalized_limit)
    else:
        pending = [case for case in cases if not case.get("result")]
        selected = _slice_cases(pending or cases, normalized_limit)
    return {
        "name": path.name,
        "path": str(path),
        "system": system,
        **metadata,
        "testcase_kind": testcase_kind,
        "total": len(cases),
        "pending": len(pending),
        "safe_total": len([case for case in cases if case.get("risk") == "SAFE"]),
        "selected": selected,
        "selected_count": len(selected),
        "limit": "all" if normalized_limit is None else normalized_limit,
        "used_pending": bool(pending),
    }


def parse_testcase_result_lines(output: str) -> list[dict[str, Any]]:
    results: dict[int, dict[str, Any]] = {}
    pattern = re.compile(
        r"RESULT\s*[:：]\s*TC\s*0*(\d+)\s*[|｜]\s*"
        r"(PASS|FAIL|N/A|NA)\s*(?:[|｜]\s*(.*))?",
        re.IGNORECASE,
    )
    for line in str(output or "").splitlines():
        match = pattern.search(line)
        if not match:
            continue
        no = int(match.group(1))
        status = match.group(2).upper().replace("NA", "N/A")
        evidence = (match.group(3) or "").strip()
        results[no] = {
            "no": no,
            "status": status,
            "evidence": evidence[:1000],
            "raw": line.strip()[:1200],
        }
    return [results[key] for key in sorted(results)]


def write_testcase_results_from_output(name: str, output: str) -> dict[str, Any]:
    parsed = parse_testcase_result_lines(output)
    if not parsed:
        return {
            "ok": True,
            "updated": 0,
            "parsed": 0,
            "missing": [],
            "message": "未找到 RESULT 行，略過 Excel 回寫。",
        }
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，無法回寫 TestCase xlsx") from e

    path = testcase_path(name)
    if not path:
        raise FileNotFoundError(f"找不到 TestCase 文件：{name}")

    metadata = _read_xlsx_metadata(path)
    testcase_kind = metadata.get("testcase_kind", "standard") or "standard"
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[wb.sheetnames[0]]
        if testcase_kind == "destructive":
            tc_col, result_col, note_col = 4, 5, 7
        else:
            tc_col, result_col, note_col = 3, 4, 6
        row_by_no: dict[int, int] = {}
        case_no = 0
        for row_index in range(2, ws.max_row + 1):
            tc_text = str(ws.cell(row_index, tc_col).value or "").strip()
            if not tc_text:
                continue
            case_no += 1
            row_by_no[case_no] = row_index

        updated = 0
        missing: list[int] = []
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        for item in parsed:
            row_index = row_by_no.get(int(item["no"]))
            if not row_index:
                missing.append(int(item["no"]))
                continue
            ws.cell(row_index, result_col).value = item["status"]
            evidence = item.get("evidence") or item.get("raw") or ""
            if evidence:
                note_cell = ws.cell(row_index, note_col)
                old_note = str(note_cell.value or "").strip()
                new_note = f"[{stamp} Agent] {evidence}"
                note_cell.value = f"{old_note}\n{new_note}" if old_note else new_note
            updated += 1
        wb.save(path)
    except PermissionError as e:
        return {
            "ok": False,
            "updated": 0,
            "parsed": len(parsed),
            "missing": [],
            "path": str(path),
            "error": f"Excel 檔可能正在開啟，無法回寫：{e}",
        }
    finally:
        wb.close()
    return {
        "ok": True,
        "updated": updated,
        "parsed": len(parsed),
        "missing": missing,
        "path": str(path),
        "message": f"已回寫 {updated}/{len(parsed)} 條 TestCase 結果。",
    }


def build_agent_task_from_testcase(name: str, game: dict,
                                   limit: int | str | None = 25) -> dict[str, Any]:
    meta = read_testcase_cases(name, limit=limit)
    if not meta["selected"]:
        if meta.get("testcase_kind") == "destructive":
            raise ValueError(f"破壞性 TestCase 沒有 SAFE 可執行案例：{name}")
        raise ValueError(f"TestCase 文件沒有可執行案例：{name}")
    if meta.get("testcase_kind") == "destructive":
        lines = "\n".join(
            f"- TC{case['no']:03d}｜{case['item'] or '未分類'}｜"
            f"{case.get('risk', 'SAFE')}｜{case['type'] or '未分類'}｜{case['tc']}"
            for case in meta["selected"]
        )
        scope = "SAFE 且未填 PASS/FAIL 的破壞性案例" if meta["used_pending"] else "SAFE 破壞性案例"
        role_line = "請以 QA 測試員身份執行破壞性 TestCase 文件中的 SAFE 測項。"
        destructive_rules = """- 這是破壞性測試，但本次只允許執行列出的 SAFE 案例。
- 嚴禁自行補做 GUARDED 或 MANUAL 案例；遇到任何可能消耗資源、變更帳號、購買、抽卡、PVP、刪除資料或不可逆狀態的畫面，立刻停止並回報 N/A。
- 若 SAFE 案例因畫面狀態變成高風險，停止該案例，不要硬做。"""
    else:
        lines = "\n".join(
            f"- TC{case['no']:03d}｜{case['item'] or '未分類'}｜"
            f"{case['type'] or '未分類'}｜{case['tc']}"
            for case in meta["selected"]
        )
        scope = "未填 PASS/FAIL 的案例" if meta["used_pending"] else "全部案例"
        role_line = "請以 QA 測試員身份執行 TestCase 文件中的測項。"
        destructive_rules = ""
    game_name = game.get("name") or game.get("id") or "指定遊戲"
    system_skill_path = str(meta.get("system_skill_path", "") or "").strip()
    system_skill_line = (
        f"系統理解 Skill：{system_skill_path}（開始測試前先讀取，理解用途、入口、規則與風險）\n"
        if system_skill_path else ""
    )
    task = f"""{role_line}

TestCase 文件：{meta['name']}
遊戲：{game_name}
目標系統/功能：{meta['system']}
{system_skill_line}來源企劃書：{meta.get('source_doc', '') or '未記錄'}
本次範圍：{scope}，先測最多 {meta['selected_count']} / {meta['total']} 條。

請先啟動或切回遊戲，導航到「{meta['system']}」相關介面，再逐條驗證下列 TestCase：
{lines}

執行規則：
- 每條 TestCase 操作前後都要截圖判讀，不要盲點。
{destructive_rules}
- 若有系統理解 Skill，先用它理解該系統要做什麼、入口在哪、有哪些狀態與風險；TestCase 仍是 PASS/FAIL 判定來源。
- 可以用遊戲 skill、圖片記憶與目前畫面判斷入口位置。
- 登入、付費、消費、抽卡確認、第三方授權、PVP 排位都不可代操作；遇到就停止該案例並標 N/A 或 FAIL，說明原因。
- 對每條案例輸出：RESULT: TC編號|PASS/FAIL/N/A|看到的證據或原因；系統會用這個格式回寫 Excel。
- 所有列出的案例都有 RESULT 後，輸出 SUITE DONE 並結束任務。
"""
    meta["task"] = task
    return meta


def _slugify_ascii(value: str, fallback: str = "skill") -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", str(value or "").lower()).strip("-")
    return slug or fallback


def _yaml_quote(value: str) -> str:
    text = str(value or "").replace("\\", "\\\\").replace('"', '\\"')
    return f'"{text}"'


def _one_line(value: str, limit: int = 240) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:limit]


def _system_skill_name(game_id: str, source_name: str, system: str) -> str:
    game_slug = _slugify_ascii(game_id, "game")[:36].strip("-") or "game"
    digest = hashlib.sha1(
        f"{game_id}:{source_name}:{system}".encode("utf-8")
    ).hexdigest()[:10]
    return f"{game_slug}-system-{digest}"[:63].rstrip("-")


def _update_system_skill_index(game_id: str, item: dict[str, Any]) -> Path:
    game_dir = SYSTEM_SKILLS_DIR / _slugify_ascii(game_id, "game")
    index_path = game_dir / "index.json"
    try:
        data = json.loads(index_path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            data = {}
    except (OSError, json.JSONDecodeError):
        data = {}
    systems = data.get("systems")
    if not isinstance(systems, dict):
        systems = {}
    systems[str(item["name"])] = {
        "name": item["name"],
        "game_id": item["game_id"],
        "system": item["system"],
        "source_doc": item["source_doc"],
        "path": item["relative_path"],
        "testcase_count": item["testcase_count"],
        "updated_at": item["updated_at"],
    }
    data = {
        "game_id": game_id,
        "updated_at": item["updated_at"],
        "systems": dict(sorted(systems.items())),
    }
    game_dir.mkdir(parents=True, exist_ok=True)
    index_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return index_path


def write_system_skill_from_testcase(
    game: dict | None,
    source_name: str,
    system: str,
    cases: list[tuple],
    issues: list[str],
    skill_notes: list[str] | None = None,
) -> dict[str, Any]:
    if not game:
        return {"updated": False, "reason": "no game"}
    game_id = str(game.get("id") or "").strip()
    if not game_id:
        return {"updated": False, "reason": "no game_id"}

    system_name = system or "未命名系統"
    game_name = game.get("name") or game_id
    skill_name = _system_skill_name(game_id, source_name, system_name)
    game_dir = SYSTEM_SKILLS_DIR / _slugify_ascii(game_id, "game")
    skill_dir = game_dir / skill_name
    skill_path = skill_dir / "SKILL.md"
    created = not skill_path.exists()
    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    counts: dict[str, int] = {}
    samples: list[str] = []
    for item, typ, priority, automation, tc in cases:
        key = item or "未分類"
        counts[key] = counts.get(key, 0) + 1
        if len(samples) < 12:
            samples.append(
                f"- {key}｜{typ}｜{priority}/{automation}｜{_one_line(tc, 180)}"
            )

    note_lines = "\n".join(
        f"- {_one_line(note, 220)}"
        for note in (skill_notes or [])[:10]
        if _one_line(note)
    )
    if not note_lines:
        note_lines = (
            "- 以來源企劃書與 TestCase 為準理解此系統；未明確描述的入口、數值或流程不要自行推測。"
        )
    count_lines = "\n".join(
        f"- {item}: {count} 條 TestCase" for item, count in counts.items()
    ) or "- 無"
    sample_lines = "\n".join(samples) if samples else "- 無"
    issue_lines = "\n".join(
        f"- {_one_line(issue, 220)}" for issue in issues[:12]
    ) if issues else "- 無"

    description = (
        f"Understand and test the {game_name} {system_name} system from "
        f"planning document {source_name}. Use when running AutoGameTest QA "
        "TestCase jobs, navigating this feature, or interpreting its intended "
        "behavior, states, risks, and acceptance criteria."
    )
    body = f"""---
name: {skill_name}
description: {_yaml_quote(description)}
---

# {game_name} / {system_name}

## System Intent

{note_lines}

## Source

- Game: {game_name} (`{game_id}`)
- Planning doc: {source_name}
- Generated: {updated_at[:10]}
- TestCase count: {len(cases)}

## Functional Areas

{count_lines}

## Agent Guidance

- Read this skill before executing QA TestCase jobs for this system.
- Use it to understand what the system is for, where to navigate, which states matter, and which risks should stop automation.
- Treat the generated TestCase workbook as the source of truth for PASS/FAIL; this skill provides context, not permission to invent missing rules.
- Stop and report if the test reaches login, account binding, payment, purchase confirmation, gacha confirmation, third-party authorization, PVP, or any unclear high-risk screen.

## Representative TestCases

{sample_lines}

## Open Questions

{issue_lines}
"""
    skill_dir.mkdir(parents=True, exist_ok=True)
    skill_path.write_text(body, encoding="utf-8")
    relative_path = os.path.relpath(skill_path, ROOT).replace(os.sep, "/")
    index_path = _update_system_skill_index(game_id, {
        "name": skill_name,
        "game_id": game_id,
        "system": system_name,
        "source_doc": source_name,
        "relative_path": relative_path,
        "testcase_count": len(cases),
        "updated_at": updated_at,
    })
    return {
        "updated": True,
        "created": created,
        "game_id": game_id,
        "system": system_name,
        "name": skill_name,
        "path": str(skill_path.resolve()),
        "relative_path": relative_path,
        "index_path": str(index_path.resolve()),
        "index_relative_path": os.path.relpath(index_path, ROOT).replace(os.sep, "/"),
    }


def autopush_files(paths: list[Path], message: str) -> str:
    files = [os.path.relpath(p, ROOT) for p in paths if p and p.exists()]
    if not files:
        return "no files"
    try:
        subprocess.run(["git", "add", "--", *files], cwd=ROOT, check=True,
                       capture_output=True, text=True, encoding="utf-8", errors="replace")
        diff = subprocess.run(["git", "diff", "--cached", "--quiet", "--", *files],
                              cwd=ROOT)
        if diff.returncode == 0:
            return "no changes"
        commit = subprocess.run(["git", "commit", "-m", message, "--", *files],
                                cwd=ROOT,
                                capture_output=True, text=True, encoding="utf-8",
                                errors="replace", timeout=120)
        if commit.returncode != 0:
            return (commit.stderr or commit.stdout or "git commit failed").strip()
        push = subprocess.run(["git", "push"], cwd=ROOT,
                              capture_output=True, text=True, encoding="utf-8",
                              errors="replace", timeout=180)
        if push.returncode != 0:
            return f"commit ok, push failed: {(push.stderr or push.stdout).strip()}"
        return "committed and pushed"
    except Exception as e:
        return f"git failed: {e}"


def generate_destructive_testcases(testcase_name: str, run_ai,
                                   on_progress=None,
                                   autopush: bool = True) -> dict[str, Any]:
    try:
        meta = read_testcase_cases(testcase_name, limit=80)
    except Exception as e:
        return {"ok": False, "error": str(e)}
    if meta.get("testcase_kind") == "destructive":
        return {"ok": False, "error": "請選擇標準 TestCase 來產生破壞性測試"}
    if not meta.get("selected"):
        return {"ok": False, "error": f"TestCase 沒有可參考案例：{testcase_name}"}

    system_skill = ""
    skill_path = _resolve_project_path(meta.get("system_skill_path", ""))
    if skill_path and skill_path.is_file() and _path_is_under(skill_path, ROOT):
        system_skill = skill_path.read_text(encoding="utf-8", errors="replace")

    if on_progress:
        on_progress("Codex 生成破壞性 TestCase 中...")
    prompt = build_destructive_prompt(meta, system_skill)
    ai_result = run_ai(prompt)
    output = ai_result.get("output", "") if isinstance(ai_result, dict) else ""
    attempts = ai_result.get("attempts", []) if isinstance(ai_result, dict) else []
    log_path = INPUT_DIR / f"{_safe_stem(testcase_name)}_destructive_{datetime.now():%Y%m%d_%H%M%S}.log"
    ensure_dirs()
    log_path.write_text(output, encoding="utf-8", errors="replace")
    parsed = parse_destructive_output(output)
    if not parsed["cases"]:
        tail = "\n".join(output.strip().splitlines()[-8:])
        return {
            "ok": False,
            "error": f"AI 未輸出任何破壞性測試案例。\n{tail}",
            "attempts": attempts,
            "log": str(log_path),
        }

    xlsx = write_destructive_xlsx(
        parsed["system"] or meta.get("system") or Path(testcase_name).stem,
        parsed["cases"],
        parsed["issues"],
        meta["name"],
        meta,
    )
    risk_counts = {
        "SAFE": sum(1 for case in parsed["cases"] if case[1] == "SAFE"),
        "GUARDED": sum(1 for case in parsed["cases"] if case[1] == "GUARDED"),
        "MANUAL": sum(1 for case in parsed["cases"] if case[1] == "MANUAL"),
    }
    git = ""
    if autopush:
        git = autopush_files(
            [xlsx],
            f"[Hibari] 新增破壞性 TestCase {xlsx.name}",
        )
    return {
        "ok": True,
        "xlsx": str(xlsx),
        "xlsx_name": xlsx.name,
        "source_testcase": meta["name"],
        "cases": len(parsed["cases"]),
        "issues": len(parsed["issues"]),
        "risk_counts": risk_counts,
        "git": git,
        "attempts": attempts,
        "log": str(log_path),
        "message": (
            f"已生成 {len(parsed['cases'])} 條破壞性 TestCase"
            f"（SAFE {risk_counts['SAFE']} / GUARDED {risk_counts['GUARDED']} / MANUAL {risk_counts['MANUAL']}）"
            + (f"；git：{git}" if git else "")
        ),
    }


def generate_testcases(doc_path: str, run_ai, on_progress=None,
                       doc_name: str | None = None,
                       game: dict | None = None,
                       autopush: bool = True) -> dict[str, Any]:
    doc = Path(doc_path)
    if not doc.exists():
        return {"ok": False, "error": f"找不到企劃書：{doc_path}"}
    if doc.suffix.lower() not in SUPPORTED_EXTS:
        return {"ok": False, "error": f"不支援的企劃書格式：{doc.suffix}"}
    ensure_dirs()
    if on_progress:
        on_progress("抽取企劃書內容...")
    text, mode = extract_doc_text(doc)
    if not text.strip() or mode == "pdf-unreadable":
        return {"ok": False, "error": text or "企劃書沒有可讀文字"}
    source_name = _safe_name(doc_name or doc.name)
    prompt = build_prompt(source_name, text, mode)
    if on_progress:
        on_progress("Codex 生成 QA TestCase 中...")
    ai_result = run_ai(prompt)
    output = ai_result.get("output", "") if isinstance(ai_result, dict) else ""
    attempts = ai_result.get("attempts", []) if isinstance(ai_result, dict) else []
    log_path = INPUT_DIR / f"{_safe_stem(source_name)}_gen_{datetime.now():%Y%m%d_%H%M%S}.log"
    log_path.write_text(output, encoding="utf-8", errors="replace")
    parsed = parse_output(output)
    if not parsed["cases"]:
        tail = "\n".join(output.strip().splitlines()[-8:])
        return {
            "ok": False,
            "error": f"AI 未輸出任何測試案例。\n{tail}",
            "attempts": attempts,
            "log": str(log_path),
        }
    system_name = parsed["system"] or doc.stem
    system_skill = write_system_skill_from_testcase(
        game,
        source_name,
        system_name,
        parsed["cases"],
        parsed["issues"],
        parsed.get("skill_notes") or [],
    ) if game else {"updated": False, "reason": "no game"}
    xlsx = write_testcase_xlsx(
        system_name,
        parsed["cases"],
        parsed["issues"],
        source_name,
        game=game,
        system_skill=system_skill if system_skill.get("updated") else None,
    )
    doc_backup = TESTCASE_DIR / source_name
    try:
        if doc.resolve() != doc_backup.resolve():
            shutil.copy2(doc, doc_backup)
    except OSError:
        doc_backup = None
    git = ""
    if autopush:
        push_paths = [xlsx] + ([doc_backup] if doc_backup else [])
        if system_skill.get("updated") and system_skill.get("path"):
            push_paths.append(Path(system_skill["path"]))
        if system_skill.get("updated") and system_skill.get("index_path"):
            push_paths.append(Path(system_skill["index_path"]))
        git = autopush_files(
            push_paths,
            f"[Hibari] 新增 QA TestCase {xlsx.name}",
        )
    return {
        "ok": True,
        "xlsx": str(xlsx),
        "xlsx_name": xlsx.name,
        "source_name": source_name,
        "doc_backup": str(doc_backup) if doc_backup else "",
        "cases": len(parsed["cases"]),
        "issues": len(parsed["issues"]),
        "mode": mode,
        "game_id": (game or {}).get("id", ""),
        "system_skill": system_skill,
        "skill_update": system_skill,
        "git": git,
        "attempts": attempts,
        "log": str(log_path),
        "message": (
            f"已生成 {len(parsed['cases'])} 條 TestCase"
            + (f"，{len(parsed['issues'])} 個待釐清問題" if parsed["issues"] else "")
            + ("，已建立/更新系統 Skill" if system_skill.get("updated") else "")
            + (f"；git：{git}" if git else "")
        ),
    }
